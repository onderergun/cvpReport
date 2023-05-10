# Copyright (c) 2021 Arista Networks, Inc.
# Use of this source code is governed by the Apache License 2.0
# that can be found in the COPYING file.
from datetime import datetime, timedelta
from google.protobuf.timestamp_pb2 import Timestamp
from cloudvision.Connector.codec.custom_types import FrozenDict
from cloudvision.Connector.codec import Wildcard, Path
from cloudvision.Connector.grpc_client import GRPCClient, create_query
from parser import base
import time

from cvprac.cvp_client import CvpClient
from cvprac.cvp_api import CvpApi
import urllib3
import ssl

import openpyxl
from openpyxl.styles import Font

import smtplib
import os.path as op
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from email import encoders

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
ssl._create_default_https_context = ssl._create_unverified_context

def send_mail(send_from, send_to, subject, message, files=[],
              server="localhost", port=587, username='', password='',
              use_tls=False):
    """Compose and send email with provided info and attachments.

    Args:
        send_from (str): from name
        send_to (str): to name
        subject (str): message title
        message (str): message body
        files (list[str]): list of file paths to be attached to email
        server (str): mail server host name
        port (int): port number
        username (str): server auth username
        password (str): server auth password
        use_tls (bool): use TLS mode
    """
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = COMMASPACE.join(send_to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(message))

    for path in files:
        part = MIMEBase('application', "octet-stream")
        with open(path, 'rb') as file:
            part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        'attachment; filename="{}"'.format(op.basename(path)))
        msg.attach(part)

    smtp = smtplib.SMTP(server, port)
    if use_tls:
        smtp.starttls()
    #smtp.login(username, password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()

def main(apiserverAddr=None, dId=None, token=None, cert=None, key=None, ca=None, days=0, hours=24, minutes=0):

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    ssl._create_default_https_context = ssl._create_unverified_context
    
    d1 = time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())
    currenttime=time.time()

    startDtime = datetime.now() - timedelta(days=days, hours=hours, minutes=minutes)
    endDtime = datetime.now()
    start = Timestamp()
    end = Timestamp()
    start.FromDatetime(startDtime)  # type: ignore
    end.FromDatetime(endDtime)  # type: ignore

    tok = "Enter your token here"
    clnt = CvpClient()
    clnt.connect(nodes=["Enter your CVP IP Address"], username='',password='',api_token=tok)
    clntapi = CvpApi(clnt)
    
    inventoryList = clntapi.get_inventory()
    
    numDevices = len(inventoryList)
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']

    sheet['A1'] = 'Hostname'
    sheet['A1'].font = Font(size=14, bold=True)
    sheet.column_dimensions['A'].width = 20
    sheet['B1'] = 'Model'
    sheet['B1'].font = Font(size=14, bold=True)
    sheet.column_dimensions['B'].width = 20
    sheet['C1'] = 'SW Version'
    sheet['C1'].font = Font(size=14, bold=True)
    sheet.column_dimensions['C'].width = 20
    sheet['D1'] = 'IP Address'
    sheet['D1'].font = Font(size=14, bold=True)
    sheet.column_dimensions['D'].width = 15
    sheet['E1'] = 'Serial Number'
    sheet['E1'].font = Font(size=14, bold=True)
    sheet.column_dimensions['E'].width = 20
    sheet['F1'] = 'Daily Availability (%)'
    sheet['F1'].font = Font(size=14, bold=True)
    sheet.column_dimensions['G'].width = 25
    sheet['G1'] = 'CPU Utilization (%)'
    sheet['G1'].font = Font(size=14, bold=True)
    sheet.column_dimensions['H'].width = 15
    sheet['H1'] = 'Memory Utilization(%)'
    sheet['H1'].font = Font(size=14, bold=True)
    sheet.column_dimensions['I'].width = 20

    k=0
    for device in inventoryList:
        sheet.cell(row=k+2,column=1).value=device["hostname"]
        sheet.cell(row=k+2,column=2).value=device["modelName"]
        sheet.cell(row=k+2,column=3).value=device["version"]
        sheet.cell(row=k+2,column=4).value=device["ipAddress"]
        sheet.cell(row=k+2,column=5).value=device["serialNumber"]
        
        pathElts = [
            "Kernel",
            "proc",
            "cpu",
            "utilization",
            "total"
        ]
        dataset = device["serialNumber"]
        query = [create_query([(pathElts, ["util"])], dataset)
                 ]
        datapoints = 0
        cpuTotal = 0
        with GRPCClient(apiserverAddr, token=token, certs=cert, key=key,
                        ca=ca) as client:
            for batch in client.get(query, start=start, end=end):
                for notif in batch["notifications"]:
                    cpuTotal += notif["updates"]["util"]
                datapoints = datapoints + len(batch["notifications"])
            sheet.cell(row=k+2,column=7).value = f"{cpuTotal/datapoints:.2f}"
    
        pathElts = [
            "Kernel",
            "proc",
            "meminfo"
        ]
        query = [create_query([(pathElts, ["memAvailable","memTotal"])], dataset)
                 ]
        datapoints = 0
        memAvailable = 0
        with GRPCClient(apiserverAddr, token=token, certs=cert, key=key,
                        ca=ca) as client:
            for batch in client.get(query, start=start, end=end):
                for notif in batch["notifications"]:
                    if "memTotal" in notif["updates"]:
                        totalMemory = notif["updates"]["memTotal"]
                    if "memAvailable" in notif["updates"]:
                        memAvailable += notif["updates"]["memAvailable"]
                datapoints = datapoints + len(batch["notifications"])
            sheet.cell(row=k+2,column=8).value = f"{100*(1-memAvailable/(totalMemory*datapoints)):.2f}"
        
        # Estimates daily uptime dramatically less than the actual.
        pathElts = [
            "Kernel",
            "sysinfo"
        ]
        query = [create_query([(pathElts, ["uptime"])], dataset)
                 ]
        uptime = 86400
        uptimelast = 0
    
        with GRPCClient(apiserverAddr, token=token, certs=cert, key=key,
                        ca=ca) as client:
            for batch in client.get(query, start=start, end=end):
                for notif in batch["notifications"]:
                    if notif["updates"]["uptime"] < uptime:
                        uptime = notif["updates"]["uptime"]
                        if uptimelast < uptime:
                            uptimelast = uptime
    
            sheet.cell(row=k+2,column=6).value = 100*(1 - uptimelast/86400)
        
        k=k+1

    #    Get the completed tasks and who executed it for the last 7 days"
    last7days = []
    for i in range (7):
        last7days.append((datetime.now() - timedelta(days=i)).strftime('%d-%m-%Y'))
    
    sheet['A'+str(numDevices+5)] = 'Username'
    sheet['A'+str(numDevices+5)].font = Font(size=14, bold=True)
    sheet['B'+str(numDevices+4)] = '# of config changes'
    sheet['B'+str(numDevices+4)].font = Font(size=14, bold=True)
    sheet['B'+str(numDevices+5)] = last7days[0]
    sheet['B'+str(numDevices+5)].font = Font(size=14, bold=True)
    sheet['C'+str(numDevices+5)] = last7days[1]
    sheet['C'+str(numDevices+5)].font = Font(size=14, bold=True)
    sheet['D'+str(numDevices+5)] = last7days[2]
    sheet['D'+str(numDevices+5)].font = Font(size=14, bold=True)
    sheet['E'+str(numDevices+5)] = last7days[3]
    sheet['E'+str(numDevices+5)].font = Font(size=14, bold=True)
    sheet['F'+str(numDevices+5)] = last7days[4]
    sheet['F'+str(numDevices+5)].font = Font(size=14, bold=True)
    sheet['G'+str(numDevices+5)] = last7days[5]
    sheet['G'+str(numDevices+5)].font = Font(size=14, bold=True)
    sheet['H'+str(numDevices+5)] = last7days[6]
    sheet['H'+str(numDevices+5)].font = Font(size=14, bold=True)
    
    taskList = clntapi.get_tasks_by_status("COMPLETED")

    userTask = {}
    for task in taskList:
        taskCreator = task["createdBy"]
        if not taskCreator in userTask:
            userTask[taskCreator] = {}
        taskCompletiontime = datetime.fromtimestamp(task["completedOnInLongFormat"]/1000).strftime('%d-%m-%Y')
        if not taskCompletiontime in userTask[taskCreator]:
            userTask[taskCreator][taskCompletiontime] = 0
        userTask[taskCreator][taskCompletiontime] = userTask[taskCreator][taskCompletiontime]+1

    k=0
    for user in userTask:
        sheet.cell(row=numDevices+6+k,column=1).value = user
        j=2
        for day in last7days:
            if day in userTask[user]:
                sheet.cell(row=numDevices+6+k,column=j).value = userTask[user][day]
            else:
                sheet.cell(row=numDevices+6+k,column=j).value = 0
            j+=1
        k+=1

    filename = 'rapor_'+ d1 + '.xlsx'
    wb.save(filename)

    send_from = "email@domainname"
    send_to = ["targetemail@domainname"]
    subject = "Daily Report "+d1
    message = ""
    files = [filename]
    server = "Enter your email server IP"
    username = ""
    password = ""
    port = 25
    use_tls = False
    send_mail(send_from, send_to, subject, message, files,
              server, port, username, password, use_tls)

if __name__ == "__main__":
    exit(main(apiserverAddr="CVP IP Address:443",ca="cvp.crt",
              cert=None, key=None, token="token.txt"))
